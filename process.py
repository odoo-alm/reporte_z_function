# =============================================================================
# process.py — GCS → BigQuery
# Lee archivos de gs://gestion-365-data-lake, parsea e inserta en BQ.
# Endpoint: POST /process
# Body: {"fecha": "2026/03/29", "etiqueta": "BARAKO", "tipo": "REPZ"}
#       Todos los parámetros son opcionales — sin ellos procesa todo el bucket.
# =============================================================================

import io
import json
import logging
import re
from datetime import datetime, timezone
from flask import Blueprint, jsonify, request
from google.cloud import bigquery, storage
import openpyxl

logger = logging.getLogger(__name__)

BQ_PROJECT = "gestion-365"
GCS_BUCKET = "gestion-365-data-lake"

BQ_TABLE_Z    = "mailer_raw.mailer_z_received"
BQ_TABLE_FACT = "mailer_raw.mailer_fact_received"
BQ_TABLE_NC   = "mailer_raw.mailer_nc_received"
BQ_TABLE_UBII = "mailer_raw.mailer_ubii_raw"
BQ_TABLE_PDYA = "mailer_raw.mailer_pdya_raw"
BQ_TABLE_MEGA      = "mailer_raw.mailer_megasoft_raw"
BQ_TABLE_CREDICARD = "mailer_raw.mailer_credicard_raw"

MEDIOS_CONOCIDOS = sorted([
    "MEGA - MAN", "MEGA", "UBII", "EFEC - DOLAR", "EFEC - BS",
    "ZELLE", "CXC", "BANESCO - PM", "BANESCO - DEB",
    "BANPLUS - CRED", "BANPLUS - DEB", "MERCAN - PM", "SALDO",
], key=len, reverse=True)

# ---------------------------------------------------------------------------
# Blueprint
# ---------------------------------------------------------------------------
process_bp = Blueprint("process", __name__)
_bq  = None
_gcs = None

def init_process(bq_client, gcs_client):
    global _bq, _gcs
    _bq  = bq_client
    _gcs = gcs_client

# ---------------------------------------------------------------------------
# Clasificador
# ---------------------------------------------------------------------------
def classify(filename):
    fu = filename.upper()
    if fu.startswith("REPZ")         and fu.endswith(".TXT"):  return "REPZ"
    if fu.startswith("FAC")          and fu.endswith(".TXT"):  return "FACT"
    if (fu.startswith("NDC") or fu.startswith("NC")) and fu.endswith(".TXT"):  return "NC"
    if "LIQUIDACION"   in fu         and (fu.endswith(".XLSX") or fu.endswith(".XLS")): return "UBII"
    if "ORDERDETAILS"  in fu         and (fu.endswith(".XLSX") or fu.endswith(".XLS")): return "PDYA"
    if "TRANSACCIONES" in fu         and (fu.endswith(".XLSX") or fu.endswith(".XLS")): return "MEGA"
    if fu.endswith(".JSON"):                                                             return "CREDICARD"
    return None

# ---------------------------------------------------------------------------
# Utilidades TXT
# ---------------------------------------------------------------------------
def _search(pattern, text, group=1):
    m = re.search(pattern, text, re.MULTILINE)
    return m.group(group).strip() if m else None

def _parse_bs(raw):
    if raw is None:
        return None
    try:
        return float(str(raw).replace(".", "").replace(",", "."))
    except ValueError:
        return None

def _bs(pattern, text):
    return _parse_bs(_search(pattern, text))

def _extract_doc_datetime(doc_label, content):
    pat = rf"{re.escape(doc_label)}\s+\d+\s*\r?\nFECHA:\s+(\d{{2}}-\d{{2}}-\d{{4}})\s+HORA:\s+(\d{{2}}:\d{{2}})"
    m = re.search(pat, content)
    return (m.group(1), m.group(2)) if m else (None, None)

def _extract_section(content, title, next_titles):
    start_m = re.search(rf"^\s*{re.escape(title)}\s*$", content, re.MULTILINE | re.IGNORECASE)
    if not start_m:
        return ""
    start_pos = start_m.end()
    end_pos = len(content)
    for nt in next_titles:
        end_m = re.search(rf"^\s*{re.escape(nt)}\s*$", content[start_pos:], re.MULTILINE | re.IGNORECASE)
        if end_m:
            end_pos = min(end_pos, start_pos + end_m.start())
    return content[start_pos:end_pos]

def _bs_in_section(block, label):
    return _parse_bs(_search(rf"^\s*{re.escape(label)}\s+Bs\s*([\-\d.,]+)", block))

def _parse_section(block, prefix=""):
    p = prefix
    return {
        f"{p}exento":    _bs_in_section(block, f"{p}EXENTO"),
        f"{p}bi_g16":    _bs_in_section(block, f"{p}BI G (16,00%)"),
        f"{p}iva_g16":   _bs_in_section(block, f"{p}IVA G (16,00%)"),
        f"{p}bi_r8":     _bs_in_section(block, f"{p}BI R (8,00%)"),
        f"{p}iva_r8":    _bs_in_section(block, f"{p}IVA R (8,00%)"),
        f"{p}bi_a31":    _bs_in_section(block, f"{p}BI A (31,00%)"),
        f"{p}iva_a31":   _bs_in_section(block, f"{p}IVA A (31,00%)"),
        f"{p}percibido": _bs_in_section(block, f"{p}PERCIBIDO"),
    }

def _extract_medios_pago(block):
    medios = {}
    for m in re.finditer(r"^(.+?)\s*\(#(\d+)\)\s+Bs\s*([\-\d.,]+)", block, re.MULTILINE):
        medios[m.group(1).strip()] = {"count": int(m.group(2)), "amount": _parse_bs(m.group(3))}
    return str(medios)

def _extract_medios_pago_fact(block):
    medios = []
    for line in block.split("\n"):
        ls = line.strip()
        if not ls:
            continue
        for medio in MEDIOS_CONOCIDOS:
            if ls.startswith(medio):
                m = re.match(rf"^{re.escape(medio)}\s+Bs\s*([\-\d.,]+)", ls)
                if m:
                    medios.append({"medio": medio, "monto": _parse_bs(m.group(1))})
                    break
    return str(medios)

# ---------------------------------------------------------------------------
# Parser: Reporte Z
# Campos BQ (111): alineados con INFORMATION_SCHEMA
# ---------------------------------------------------------------------------
def parse_repz(content, filename, etiqueta):
    if not re.search(r"REPORTE Z:", content):
        return None

    id_reporte_z     = _search(r"REPORTE Z:\s+(\d+)", content)
    serial_impresora = _search(r"^MH\s+(\S+)", content)
    fecha_raw        = _search(r"FECHA:\s+(\d{2}-\d{2}-\d{4})", content)
    hora_raw         = _search(r"HORA:\s+(\d{2}:\d{2})", content)

    if not all([id_reporte_z, serial_impresora, fecha_raw]):
        return None
    try:
        fecha_reporte = datetime.strptime(fecha_raw, "%d-%m-%Y").strftime("%Y-%m-%d")
    except ValueError:
        return None

    num_facturas = _search(r"#FACT DEL DIA\s+(\d+)", content)
    num_dnf      = _search(r"#DNF DEL DIA\s+(\d+)", content)
    num_nc       = _search(r"#NC DEL DIA\s+(\d+)", content)

    f_uf,   h_uf   = _extract_doc_datetime("ULTIMA FACTURA",   content)
    f_und,  h_und  = _extract_doc_datetime("ULT.NOTA.DEBITO",  content)
    f_unc,  h_unc  = _extract_doc_datetime("ULT.NOTA.CREDITO", content)
    f_udnf, h_udnf = _extract_doc_datetime("ULTIMO DNF",       content)
    f_urmf, h_urmf = _extract_doc_datetime("ULTIMO RMF",       content)

    sec_recargos     = _extract_section(content,            "RECARGOS",         ["DESCUENTOS"])
    sec_descuentos   = _extract_section(content,            "DESCUENTOS",       ["ANULACIONES"])
    sec_anulaciones  = _extract_section(content,            "ANULACIONES",      ["CORRECCIONES"])
    sec_correcciones = _extract_section(content,            "CORRECCIONES",     ["VENTAS"])
    pos_corr         = content.find("CORRECCIONES")
    sec_ventas       = _extract_section(content[pos_corr:], "VENTAS",           ["NOTAS DE DEBITO"]) if pos_corr >= 0 else ""
    sec_nd           = _extract_section(content,            "NOTAS DE DEBITO",  ["NOTAS DE CREDITO"])
    sec_nc           = _extract_section(content,            "NOTAS DE CREDITO", ["ULTIMA FACTURA"])

    d_rec  = _parse_section(sec_recargos)
    d_desc = _parse_section(sec_descuentos)
    d_anul = _parse_section(sec_anulaciones)
    d_corr = _parse_section(sec_correcciones)
    d_vent = _parse_section(sec_ventas)
    d_nd   = _parse_section(sec_nd, prefix="ND.")
    d_nc   = _parse_section(sec_nc, prefix="NC.")

    return {
        # Posiciones 1-8
        "id_reporte_z":    id_reporte_z,
        "serial_impresora":serial_impresora,
        "fecha_reporte":   fecha_reporte,
        "hora_reporte":    hora_raw,
        "total_gaveta":    _bs(r"TOTAL GAVETA\s+Bs\s*([\-\d.,]+)", content),
        "num_facturas":    int(num_facturas) if num_facturas else None,
        "num_dnf":         int(num_dnf)      if num_dnf      else None,
        "num_nc":          int(num_nc)        if num_nc       else None,
        # Posiciones 9-17 (totales descuentos/anulaciones/correcciones)
        "subtotal_descuentos":  _bs_in_section(sec_descuentos,   "SUBTTL DESCUENTOS"),
        "iva_descuentos":       _bs_in_section(sec_descuentos,   "IVA DESCUENTOS"),
        "total_descuentos":     _bs_in_section(sec_descuentos,   "TOTAL DESCUENTOS"),
        "subtotal_anulaciones": _bs_in_section(sec_anulaciones,  "SUBTTL ANULACIONES"),
        "iva_anulaciones":      _bs_in_section(sec_anulaciones,  "IVA ANULACIONES"),
        "total_anulaciones":    _bs_in_section(sec_anulaciones,  "TOTAL ANULACIONES"),
        "subtotal_correcciones":_bs_in_section(sec_correcciones, "SUBTTL CORRECCIONES"),
        "iva_correcciones":     _bs_in_section(sec_correcciones, "IVA CORRECCIONES"),
        "total_correcciones":   _bs_in_section(sec_correcciones, "TOTAL CORRECCIONES"),
        # Posiciones 18-22 (ventas)
        "subtotal_venta":  _bs_in_section(sec_ventas, "SUBTTL VENTA"),
        "iva_venta":       _bs_in_section(sec_ventas, "IVA VENTA"),
        "igtf_venta":      _bs_in_section(sec_ventas, "IGTF VENTA (3,00%)"),
        "total_venta":     _bs_in_section(sec_ventas, "TOTAL VENTA"),
        "bi_igtf":         _bs_in_section(sec_ventas, "BI IGTF (3,00%)"),
        # Posiciones 23-26 (ND totales)
        "subtotal_nota_debito": _bs_in_section(sec_nd, "SUBTTL NOTA DEBITO"),
        "iva_nota_debito":      _bs_in_section(sec_nd, "IVA NOTA DEBITO"),
        "igtf_nota_debito":     _bs_in_section(sec_nd, "IGTF NOTA DEBITO (3,00%)"),
        "total_nota_debito":    _bs_in_section(sec_nd, "TOTAL NOTA DEBITO"),
        # Posiciones 27-30 (NC totales)
        "subtotal_nota_credito": _bs_in_section(sec_nc, "SUBTTL NOTA CREDITO"),
        "iva_nota_credito":      _bs_in_section(sec_nc, "IVA NOTA CREDITO"),
        "igtf_nota_credito":     _bs_in_section(sec_nc, "IGTF NOTA CREDITO (3,00%)"),
        "total_nota_credito":    _bs_in_section(sec_nc, "TOTAL NOTA CREDITO"),
        # Posiciones 31-34 (últimos docs — número)
        "ultima_factura":   _search(r"ULTIMA FACTURA\s+(\d+)",    content),
        "ult_nota_debito":  _search(r"ULT\.NOTA\.DEBITO\s+(\d+)", content),
        "ult_nota_credito": _search(r"ULT\.NOTA\.CREDITO\s+(\d+)",content),
        "ultimo_dnf":       _search(r"ULTIMO DNF\s+(\d+)",        content),
        # Posición 35
        "medios_pago": _extract_medios_pago(content),
        # Posiciones 36-39 (metadata)
        "etiqueta":     etiqueta,
        "filename":     filename,
        "gmail_msg_id": "",
        "processed_at": datetime.now(tz=timezone.utc).isoformat(),
        # Posiciones 40-50 (fechas/horas últimos docs)
        "fecha_ultima_factura":   f_uf,
        "hora_ultima_factura":    h_uf,
        "fecha_ult_nota_debito":  f_und,
        "hora_ult_nota_debito":   h_und,
        "fecha_ult_nota_credito": f_unc,
        "hora_ult_nota_credito":  h_unc,
        "fecha_ultimo_dnf":       f_udnf,
        "hora_ultimo_dnf":        h_udnf,
        "ultimo_rmf":             _search(r"ULTIMO RMF\s+(\d+)", content),
        "fecha_ultimo_rmf":       f_urmf,
        "hora_ultimo_rmf":        h_urmf,
        # Posiciones 51-61 (RECARGOS)
        "recargos_exento":    d_rec["exento"],
        "recargos_bi_g16":    d_rec["bi_g16"],
        "recargos_iva_g16":   d_rec["iva_g16"],
        "recargos_bi_r8":     d_rec["bi_r8"],
        "recargos_iva_r8":    d_rec["iva_r8"],
        "recargos_bi_a31":    d_rec["bi_a31"],
        "recargos_iva_a31":   d_rec["iva_a31"],
        "recargos_percibido": d_rec["percibido"],
        "subtotal_recargos":  _bs_in_section(sec_recargos, "SUBTTL RECARGOS"),
        "iva_recargos":       _bs_in_section(sec_recargos, "IVA RECARGOS"),
        "total_recargos":     _bs_in_section(sec_recargos, "TOTAL RECARGOS"),
        # Posiciones 62-69 (DESCUENTOS desglose)
        "descuentos_exento":    d_desc["exento"],
        "descuentos_bi_g16":    d_desc["bi_g16"],
        "descuentos_iva_g16":   d_desc["iva_g16"],
        "descuentos_bi_r8":     d_desc["bi_r8"],
        "descuentos_iva_r8":    d_desc["iva_r8"],
        "descuentos_bi_a31":    d_desc["bi_a31"],
        "descuentos_iva_a31":   d_desc["iva_a31"],
        "descuentos_percibido": d_desc["percibido"],
        # Posiciones 70-77 (ANULACIONES desglose)
        "anulaciones_exento":    d_anul["exento"],
        "anulaciones_bi_g16":    d_anul["bi_g16"],
        "anulaciones_iva_g16":   d_anul["iva_g16"],
        "anulaciones_bi_r8":     d_anul["bi_r8"],
        "anulaciones_iva_r8":    d_anul["iva_r8"],
        "anulaciones_bi_a31":    d_anul["bi_a31"],
        "anulaciones_iva_a31":   d_anul["iva_a31"],
        "anulaciones_percibido": d_anul["percibido"],
        # Posiciones 78-85 (CORRECCIONES desglose)
        "correcciones_exento":    d_corr["exento"],
        "correcciones_bi_g16":    d_corr["bi_g16"],
        "correcciones_iva_g16":   d_corr["iva_g16"],
        "correcciones_bi_r8":     d_corr["bi_r8"],
        "correcciones_iva_r8":    d_corr["iva_r8"],
        "correcciones_bi_a31":    d_corr["bi_a31"],
        "correcciones_iva_a31":   d_corr["iva_a31"],
        "correcciones_percibido": d_corr["percibido"],
        # Posiciones 86-93 (VENTAS desglose)
        "ventas_exento":    d_vent["exento"],
        "ventas_bi_g16":    d_vent["bi_g16"],
        "ventas_iva_g16":   d_vent["iva_g16"],
        "ventas_bi_r8":     d_vent["bi_r8"],
        "ventas_iva_r8":    d_vent["iva_r8"],
        "ventas_bi_a31":    d_vent["bi_a31"],
        "ventas_iva_a31":   d_vent["iva_a31"],
        "ventas_percibido": d_vent["percibido"],
        # Posiciones 94-102 (ND desglose)
        "nd_exento":    d_nd["ND.exento"],
        "nd_bi_g16":    d_nd["ND.bi_g16"],
        "nd_iva_g16":   d_nd["ND.iva_g16"],
        "nd_bi_r8":     d_nd["ND.bi_r8"],
        "nd_iva_r8":    d_nd["ND.iva_r8"],
        "nd_bi_a31":    d_nd["ND.bi_a31"],
        "nd_iva_a31":   d_nd["ND.iva_a31"],
        "nd_percibido": d_nd["ND.percibido"],
        "nd_bi_igtf":   _bs_in_section(sec_nd, "ND.BI IGTF (3,00%)"),
        # Posiciones 103-111 (NC desglose)
        "nc_exento":    d_nc["NC.exento"],
        "nc_bi_g16":    d_nc["NC.bi_g16"],
        "nc_iva_g16":   d_nc["NC.iva_g16"],
        "nc_bi_r8":     d_nc["NC.bi_r8"],
        "nc_iva_r8":    d_nc["NC.iva_r8"],
        "nc_bi_a31":    d_nc["NC.bi_a31"],
        "nc_iva_a31":   d_nc["NC.iva_a31"],
        "nc_percibido": d_nc["NC.percibido"],
        "nc_bi_igtf":   _bs_in_section(sec_nc, "NC.BI IGTF (3,00%)"),
    }

# ---------------------------------------------------------------------------
# Parser: Facturas y NC (mismo schema en BQ)
# ---------------------------------------------------------------------------
def _parse_doc_fiscal(block, doc_type, filename, etiqueta):
    if doc_type == "FACT":
        num = _search(r"FACTURA:\s+(\d+)", block)
    else:
        num = _search(r"NOTA DE CREDITO:\s+(\d+)", block)
    fecha_raw = _search(r"FECHA:\s+(\d{2}-\d{2}-\d{4})", block)
    if not num or not fecha_raw:
        return None
    try:
        fecha_doc = datetime.strptime(fecha_raw, "%d-%m-%Y").strftime("%Y-%m-%d")
    except ValueError:
        return None
    bi_m = re.search(
        r"BI\s*G(\d+,\d+)%\s+Bs\s*([\-\d.,]+)\s+IVA\s*G\d+,\d+%\s+Bs\s*([\-\d.,]+)",
        block
    )
    return {
        "num_documento":      num,
        "tipo_documento":     "FACTURA" if doc_type == "FACT" else "NOTA DE CREDITO",
        "num_factura_origen": _search(r"#FAC:\s+(\d+)", block) if doc_type == "NC" else None,
        "serial_control":     _search(r"#CONTROL/SERIAL IF:\s*(\S+)", block) if doc_type == "NC" else None,
        "fecha_documento":    fecha_doc,
        "hora_documento":     _search(r"HORA:\s+(\d{2}:\d{2})", block),
        "rif_cliente":        (_search(r"RIF/C\.I\.:\s*(.+)", block) or "").strip() or None,
        "razon_social":       (_search(r"RAZON SOCIAL:\s*(.+)", block) or "").strip() or None,
        "serial_impresora":   _search(r"^MH\s+(\S+)", block),
        "subtotal":           _bs(r"SUBTTL\s+Bs\s*([\-\d.,]+)", block),
        "base_imponible_16":  _parse_bs(bi_m.group(2)) if bi_m else None,
        "iva_16":             _parse_bs(bi_m.group(3)) if bi_m else None,
        "medios_pago":        _extract_medios_pago_fact(block),
        "cambio":             _bs(r"CAMBIO\s+Bs\s*([\-\d.,]+)", block),
        "igtf":               _bs(r"IGTF\d+,\d+%\s+Bs\s*([\-\d.,]+)", block),
        "total":              _bs(r"TOTAL\s+Bs\s*([\-\d.,]+)", block),
        "etiqueta":           etiqueta,
        "filename":           filename,
        "gmail_msg_id":       "",
        "processed_at":       datetime.now(tz=timezone.utc).isoformat(),
    }

def parse_facturas(content, filename, etiqueta):
    bloques = [b for b in re.split(r"(?=\s+SENIAT\s+)", content)
               if b.strip() and re.search(r"FACTURA:", b)]
    return [r for b in bloques for r in [_parse_doc_fiscal(b, "FACT", filename, etiqueta)] if r]

def parse_nc(content, filename, etiqueta):
    bloques = [b for b in re.split(r"(?=\s+SENIAT\s+)", content)
               if b.strip() and re.search(r"NOTA DE CREDITO:", b)]
    return [r for b in bloques for r in [_parse_doc_fiscal(b, "NC", filename, etiqueta)] if r]

# ---------------------------------------------------------------------------
# Parser: Ubii (XLSX)
# ---------------------------------------------------------------------------
def parse_ubii(raw_bytes, filename, etiqueta):
    sf = lambda v: float(v) if v is not None else None
    ss = lambda v: str(v).strip() if v is not None else None
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True)
    ws = wb[wb.sheetnames[0]]
    records = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or row[0] is None:
            continue
        try:
            fecha_raw = row[0]
            records.append({
                "fecha_cierre": fecha_raw.strftime("%Y-%m-%d") if hasattr(fecha_raw, "strftime") else ss(fecha_raw),
                "terminal": ss(row[1]), "oficina": ss(row[2]),
                "nro_lote": int(row[3]) if row[3] is not None else None,
                "tipo": ss(row[4]), "inmediata": ss(row[5]),
                "monto": sf(row[6]), "monto_tdc": sf(row[7]), "monto_tdd": sf(row[8]),
                "comision_tdc": sf(row[9]), "comision_islr": sf(row[10]), "comision_tdd": sf(row[11]),
                "comision_transferencia": sf(row[12]), "comision_inmediata": sf(row[13]),
                "comision_fee": sf(row[14]), "monto_cuota_credito": sf(row[15]),
                "monto_gastos_admin_credito": sf(row[16]), "monto_gastos_reactivacion_credito": sf(row[17]),
                "monto_nota_debito": sf(row[18]), "monto_ajuste_liquidacion": sf(row[19]),
                "monto_a_liquidar": sf(row[20]), "monto_tdc_visa": sf(row[21]),
                "comision_tdc_visa": sf(row[22]), "islr_tdc_visa": sf(row[23]),
                "comision_transferencia_visa": sf(row[24]), "monto_a_liquidar_visa": sf(row[25]),
                "etiqueta": etiqueta, "filename": filename,
                "gmail_msg_id": "", "processed_at": datetime.now(tz=timezone.utc).isoformat(),
            })
        except Exception as e:
            logger.warning(f"[process] ubii {filename} fila {i}: {e}")
    wb.close()
    return records

# ---------------------------------------------------------------------------
# Parser: PedidosYa (XLSX)
# ---------------------------------------------------------------------------
def parse_pedidosya(raw_bytes, filename, etiqueta):
    sf = lambda v: float(v) if v is not None and str(v).strip() != "" else None
    ss = lambda v: str(v).strip() if v is not None else None
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True)
    ws = wb[wb.sheetnames[0]]
    records = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= 1 or row[0] is None or str(row[0]).strip() == "":
            continue
        try:
            records.append({
                "nombre_local": ss(row[0]), "nro_pedido": ss(row[1]), "id_local": ss(row[2]),
                "metodo_entrega": ss(row[3]), "forma_pago": ss(row[4]),
                "pedido_programado": ss(row[5]), "direccion_restaurante": ss(row[6]),
                "estado_pedido": ss(row[7]), "fecha_pedido": ss(row[8]),
                "aceptado_en": ss(row[9]), "lista_retiro": ss(row[10]),
                "repartidor_cerca": ss(row[11]), "retirado_del_local": ss(row[12]),
                "entregado_el": ss(row[13]), "tiene_reclamos": ss(row[14]),
                "motivo_reclamo": ss(row[15]), "cancelado_el": ss(row[16]),
                "motivo_rechazo": ss(row[17]), "responsable_rechazo": ss(row[18]),
                "total_pedido": sf(row[19]), "tarifa_minima_pedido": sf(row[20]),
                "resarcimientos": sf(row[21]), "cargo_impositivo": sf(row[22]),
                "tarifa_pago": sf(row[23]), "descuentos_subsidiados_tienda": sf(row[24]),
                "voucher_subsidiado_tienda": sf(row[25]), "comision": sf(row[26]),
                "cargos": sf(row[27]), "cargos_descuentos_fugaces": sf(row[28]),
                "tarifa_servicios_publicidad": sf(row[29]),
                "es_pagadero": ss(row[30]), "pagado": ss(row[31]), "ingreso_estimado": ss(row[32]),
                "monto_efectivo_cobrado": sf(row[33]), "monto_adeudado_pedidosya": sf(row[34]),
                "monto_pago": sf(row[35]), "descuento_financiado_pedidosya": sf(row[36]),
                "voucher_financiado_pedidosya": sf(row[37]), "descuento_total": sf(row[38]),
                "total_voucher": sf(row[39]), "monto_impuestos": sf(row[40]),
                "articulos": ss(row[41]),
                "etiqueta": etiqueta, "filename": filename,
                "gmail_msg_id": "", "processed_at": datetime.now(tz=timezone.utc).isoformat(),
            })
        except Exception as e:
            logger.warning(f"[process] pdya {filename} fila {i}: {e}")
    wb.close()
    return records

# ---------------------------------------------------------------------------
# Parser: Megasoft (XLSX)
# ---------------------------------------------------------------------------
def parse_megasoft(raw_bytes, filename, etiqueta):
    si = lambda v: int(v)   if v is not None else None
    sf = lambda v: float(v) if v is not None else None
    ss = lambda v: str(v).strip() if v is not None else None
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True)
    ws = wb[wb.sheetnames[0]]
    records = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row[0] and str(row[0]).strip() == "Fecha Transacción":
                header_found = True
            continue
        if row[0] is None:
            continue
        try:
            records.append({
                "fecha_transaccion": ss(row[0]), "merchant_name": ss(row[1]),
                "afiliado": ss(row[2]), "caja": ss(row[3]), "terminal": ss(row[4]),
                "seq_num": si(row[5]), "pan_pci": ss(row[6]), "rif": ss(row[7]),
                "transaction_name": ss(row[8]), "acquirer_name": ss(row[9]),
                "card_name": ss(row[10]), "auth_id": ss(row[11]),
                "ref": ss(row[12]), "rrn": ss(row[13]),
                "fecha_cierre": ss(row[14]), "lote": ss(row[15]),
                "reversar": si(row[16]), "anulada": si(row[17]),
                "codigo_respuesta": ss(row[18]), "telefono": ss(row[19]),
                "monto": sf(row[20]), "tasa_cambio": sf(row[21]),
                "monto_cambio": sf(row[22]), "moneda_cambio": ss(row[23]),
                "etiqueta": etiqueta, "filename": filename,
                "gmail_msg_id": "", "processed_at": datetime.now(tz=timezone.utc).isoformat(),
            })
        except Exception as e:
            logger.warning(f"[process] mega {filename} fila: {e}")
    wb.close()
    return records

# ---------------------------------------------------------------------------
# Parser: Credicard (JSON)
# ---------------------------------------------------------------------------
def _parse_bs(raw):
    if not raw:
        return None
    try:
        return float(str(raw).replace(" Bs", "").replace(".", "").replace(",", ".").strip())
    except ValueError:
        return None

def parse_credicard(raw_bytes, filename, etiqueta):
    try:
        data = json.loads(raw_bytes.decode("utf-8", errors="replace"))
    except Exception as e:
        logger.error(f"[process] credicard {filename}: JSON inválido — {e}")
        return []
    records = []
    for row in data:
        if not row.get("0") or str(row.get("6", "")).strip() == "Total":
            continue
        try:
            records.append({
                "afiliado":     str(row["0"]).strip(),
                "fecha":        datetime.strptime(str(row["1"]).strip(), "%d/%m/%Y").strftime("%Y-%m-%d") if row.get("1") else None,
                "lote":         int(row["2"]) if row.get("2") else None,
                "terminal":     str(row["3"]).strip() if row.get("3") else None,
                "producto":     str(row["4"]).strip() if row.get("4") else None,
                "banco":        str(row["5"]).strip() if row.get("5") else None,
                "tarjeta":      str(row["6"]).strip() if row.get("6") else None,
                "autorizacion": str(row["7"]).strip() if row.get("7") else None,
                "monto":        _parse_bs(row.get("8")),
                "comision":     _parse_bs(row.get("9")),
                "islr":         _parse_bs(row.get("10")),
                "abonado":      _parse_bs(row.get("11")),
                "etiqueta":     etiqueta,
                "filename":     filename,
                "gmail_msg_id": "",
                "processed_at": datetime.now(tz=timezone.utc).isoformat(),
            })
        except Exception as e:
            logger.warning(f"[process] credicard {filename} fila omitida — {e}")
    logger.info(f"[process] credicard {filename}: {len(records)} filas.")
    return records

# ---------------------------------------------------------------------------
# Control de archivos ya procesados
# ---------------------------------------------------------------------------
def _get_processed_filenames(table):
    """
    Retorna el conjunto de basenames ya insertados en la tabla BQ.
    Usa basename (no el path completo) para que coincida tanto con registros
    cargados vía email-sync (guardan solo el filename) como vía GCS (guardan
    el path completo). Se llama una vez por tabla al inicio del run.
    """
    try:
        rows = _bq.query(
            f"SELECT DISTINCT REGEXP_EXTRACT(filename, r'[^/]+$') AS basename"
            f" FROM `{BQ_PROJECT}.{table}` WHERE filename IS NOT NULL"
        ).result()
        return {row.basename for row in rows if row.basename}
    except Exception as e:
        logger.warning(f"[process] No se pudo obtener filenames de {table}: {e}")
        return set()

# ---------------------------------------------------------------------------
# Carga BQ
# ---------------------------------------------------------------------------
def _load_to_bq(records, table):
    if not records:
        return 0
    buffer = io.BytesIO()
    for r in records:
        buffer.write((json.dumps(r, default=str, ensure_ascii=False) + "\n").encode())
    buffer.seek(0)
    job = _bq.load_table_from_file(
        buffer, f"{BQ_PROJECT}.{table}",
        job_config=bigquery.LoadJobConfig(
            source_format=bigquery.SourceFormat.NEWLINE_DELIMITED_JSON,
            write_disposition=bigquery.WriteDisposition.WRITE_APPEND,
            ignore_unknown_values=True,
        )
    )
    job.result()
    return job.output_rows

# ---------------------------------------------------------------------------
# Endpoint /process
# ---------------------------------------------------------------------------
@process_bp.route("/process", methods=["POST"])
def process():
    from datetime import datetime, timezone
    import uuid

    run_id    = str(uuid.uuid4())
    inicio_ts = datetime.now(tz=timezone.utc)
    body      = request.get_json(silent=True) or {}

    filtro_fecha    = body.get("fecha")     # ej: "2026/03/29"
    filtro_etiqueta = body.get("etiqueta")  # ej: "BARAKO"
    _TIPO_ALIAS = {
        "pedidosya": "PDYA", "ubii": "UBII", "megasoft": "MEGA",
        "z": "REPZ", "facturas": "FACT", "nc": "NC", "credicard": "CREDICARD",
    }
    _t = body.get("tipo")
    filtro_tipo = _TIPO_ALIAS.get(_t.lower(), _t.upper()) if _t else None

    logger.info(f"=== PROCESS run_id={run_id} | fecha={filtro_fecha} etiqueta={filtro_etiqueta} tipo={filtro_tipo} ===")

    bucket = _gcs.bucket(GCS_BUCKET)

    # Construir prefijo GCS
    prefix = ""
    if filtro_etiqueta and filtro_fecha:
        prefix = f"{filtro_etiqueta}/{filtro_fecha}/"
    elif filtro_etiqueta:
        prefix = f"{filtro_etiqueta}/"
    elif filtro_fecha:
        prefix = filtro_fecha + "/"

    blobs = list(bucket.list_blobs(prefix=prefix))
    logger.info(f"[process] {len(blobs)} archivos en GCS con prefix='{prefix}'")

    # Acumuladores por tabla
    lotes = {
        BQ_TABLE_Z: [], BQ_TABLE_FACT: [], BQ_TABLE_NC: [],
        BQ_TABLE_UBII: [], BQ_TABLE_PDYA: [], BQ_TABLE_MEGA: [],
        BQ_TABLE_CREDICARD: [],
    }
    tabla_map = {
        "REPZ": BQ_TABLE_Z, "FACT": BQ_TABLE_FACT, "NC": BQ_TABLE_NC,
        "UBII": BQ_TABLE_UBII, "PDYA": BQ_TABLE_PDYA, "MEGA": BQ_TABLE_MEGA,
        "CREDICARD": BQ_TABLE_CREDICARD,
    }

    # Cargar filenames ya procesados por tabla (una query por tabla al inicio del run)
    tablas_a_consultar = (
        [tabla_map[filtro_tipo.upper()]] if filtro_tipo and filtro_tipo.upper() in tabla_map
        else list(lotes.keys())
    )
    ya_procesados = {tabla: _get_processed_filenames(tabla) for tabla in tablas_a_consultar}
    total_skipped = sum(len(v) for v in ya_procesados.values())
    logger.info(f"[process] Filenames ya en BQ: {total_skipped} (dedup activo)")

    stats = {t.split(".")[-1]: {"archivos": 0, "registros": 0, "skipped": 0} for t in lotes}
    errores = []

    for blob in blobs:
        filename = blob.name
        basename = filename.split("/")[-1]
        # Etiqueta desde la ruta: ETIQUETA/YYYY/MM/DD/filename
        etiqueta = filename.split("/")[0].upper()

        tipo = classify(basename)
        if not tipo:
            continue
        if filtro_tipo and tipo != filtro_tipo.upper():
            continue

        # Saltar archivos ya procesados (compara por basename para cubrir
        # tanto registros cargados vía email-sync como vía GCS path completo)
        tabla = tabla_map.get(tipo)
        if tabla and basename in ya_procesados.get(tabla, set()):
            logger.info(f"[process] SKIP (ya procesado): {filename}")
            stats[tabla.split(".")[-1]]["skipped"] += 1
            continue

        try:
            raw_bytes = blob.download_as_bytes()

            if tipo in ("REPZ", "FACT", "NC"):
                content = raw_bytes.decode("utf-8", errors="replace")
                if   tipo == "REPZ": records = [r for r in [parse_repz(content, filename, etiqueta)] if r]
                elif tipo == "FACT": records = parse_facturas(content, filename, etiqueta)
                else:                records = parse_nc(content, filename, etiqueta)
            elif tipo == "UBII":      records = parse_ubii(raw_bytes, filename, etiqueta)
            elif tipo == "PDYA":      records = parse_pedidosya(raw_bytes, filename, etiqueta)
            elif tipo == "MEGA":      records = parse_megasoft(raw_bytes, filename, etiqueta)
            elif tipo == "CREDICARD": records = parse_credicard(raw_bytes, filename, etiqueta)
            else:
                continue

            lotes[tabla].extend(records)
            key = tabla.split(".")[-1]
            stats[key]["archivos"]  += 1
            stats[key]["registros"] += len(records)
            logger.info(f"[process] {tipo} {filename}: {len(records)} registros.")

        except Exception as e:
            msg = f"{filename}: {e}"
            logger.error(f"[process] ERROR {msg}")
            errores.append(msg)

    # Carga a BQ
    cargados = {}
    for tabla, records in lotes.items():
        key = tabla.split(".")[-1]
        try:
            n = _load_to_bq(records, tabla)
            cargados[key] = n
            logger.info(f"[process] {tabla}: {n} filas cargadas.")
        except Exception as e:
            msg = f"Load {tabla}: {e}"
            logger.error(f"[process] {msg}")
            errores.append(msg)
            cargados[key] = 0

    duracion = int((datetime.now(tz=timezone.utc) - inicio_ts).total_seconds())
    status   = "OK" if not errores else "PARCIAL"

    logger.info(f"=== FIN PROCESS run_id={run_id} | status={status} | {duracion}s ===")

    return jsonify({
        "run_id":    run_id,
        "status":    status,
        "duracion_s": duracion,
        "archivos":  {k: v["archivos"]  for k, v in stats.items()},
        "skipped":   {k: v["skipped"]   for k, v in stats.items()},
        "registros": {k: v["registros"] for k, v in stats.items()},
        "cargados":  cargados,
        "errores":   errores,
    }), 200 if status == "OK" else 207