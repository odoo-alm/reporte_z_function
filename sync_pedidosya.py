# =============================================================================
# sync_pedidosya.py  v3 — usa utils.py
# =============================================================================

import io
import json
import logging
import time
from datetime import datetime, timezone

import openpyxl
from google.cloud import bigquery
from utils import list_messages, get_attachments, mark_as_processed, circuit_is_open

logger = logging.getLogger(__name__)

BQ_PROJECT    = "gestion-365"
BQ_TABLE_PDYA = "mailer_raw.mailer_pdya_raw"
PAUSE_S       = 5

def parse_pedidosya(raw_bytes, filename, msg_id, etiqueta):
    sf = lambda v: float(v) if v is not None and str(v).strip() != "" else None
    ss = lambda v: str(v).strip() if v is not None else None
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True)
    ws = wb[wb.sheetnames[0]]
    records = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= 1 or row[0] is None:
            continue
        try:
            records.append({
                "nombre_local": ss(row[0]), "nro_pedido": ss(row[1]), "id_local": ss(row[2]),
                "metodo_entrega": ss(row[3]), "forma_pago": ss(row[4]),
                "pedido_programado": ss(row[5]), "direccion_restaurante": ss(row[6]),
                "estado_pedido": ss(row[7]), "fecha_pedido": ss(row[8]),
                "aceptado_en": ss(row[9]), "lista_retiro": ss(row[10]),
                "repartidor_cerca": ss(row[11]), "retirado_del_local": ss(row[12]),
                "entregado_el": ss(row[13]), "tiene_reclamos": ss(row[14]),
                "motivo_reclamo": ss(row[15]), "cancelado_el": ss(row[16]),
                "motivo_rechazo": ss(row[17]), "responsable_rechazo": ss(row[18]),
                "total_pedido": sf(row[19]), "tarifa_minima_pedido": sf(row[20]),
                "resarcimientos": sf(row[21]), "cargo_impositivo": sf(row[22]),
                "tarifa_pago": sf(row[23]), "descuentos_subsidiados_tienda": sf(row[24]),
                "voucher_subsidiado_tienda": sf(row[25]), "comision": sf(row[26]),
                "cargos": sf(row[27]), "cargos_descuentos_fugaces": sf(row[28]),
                "tarifa_servicios_publicidad": sf(row[29]),
                "es_pagadero": ss(row[30]), "pagado": ss(row[31]), "ingreso_estimado": ss(row[32]),
                "monto_efectivo_cobrado": sf(row[33]), "monto_adeudado_pedidosya": sf(row[34]),
                "monto_pago": sf(row[35]), "descuento_financiado_pedidosya": sf(row[36]),
                "voucher_financiado_pedidosya": sf(row[37]), "descuento_total": sf(row[38]),
                "total_voucher": sf(row[39]), "monto_impuestos": sf(row[40]),
                "articulos": ss(row[41]),
                "etiqueta": etiqueta, "filename": filename,
                "gmail_msg_id": msg_id, "processed_at": datetime.now(tz=timezone.utc).isoformat(),
            })
        except Exception as e:
            logger.warning(f"[sync_pedidosya] {filename}: fila {i} omitida — {e}")
    wb.close()
    logger.info(f"[sync_pedidosya] {filename}: {len(records)} filas.")
    return records

def _load_to_bq(records, bq_client):
    if not records:
        return 0
    buffer = io.BytesIO()
    for r in records:
        buffer.write((json.dumps(r, default=str, ensure_ascii=False) + "\n").encode())
    buffer.seek(0)
    job = bq_client.load_table_from_file(
        buffer, f"{BQ_PROJECT}.{BQ_TABLE_PDYA}",
        job_config=bigquery.LoadJobConfig(
            source_format=bigquery.SourceFormat.NEWLINE_DELIMITED_JSON,
            write_disposition=bigquery.WriteDisposition.WRITE_APPEND,
            ignore_unknown_values=True,
        )
    )
    job.result()
    logger.info(f"[sync_pedidosya] {job.output_rows} filas cargadas.")
    return job.output_rows

def sync_pedidosya(service, bq_client, etiquetas, label_map=None, processed_label_id=None):
    stats = {"procesados": 0, "cargados": 0, "errores": []}
    todos = []
    for etiqueta in etiquetas:
        if circuit_is_open():
            stats["errores"].append("Circuit breaker abierto — abortado.")
            break
        label_id = (label_map or {}).get(etiqueta.upper())
        if not label_id:
            continue
        messages = list_messages(service, label_id)
        logger.info(f"[sync_pedidosya][{etiqueta}] {len(messages)} correos nuevos.")
        for msg_meta in messages:
            if circuit_is_open():
                stats["errores"].append("Circuit breaker abierto — abortado.")
                break
            msg_id = msg_meta["id"]
            try:
                atts = get_attachments(service, msg_id, keyword="ORDERDETAILS", extensions=(".xlsx", ".xls"))
                for att in atts:
                    try:
                        records = parse_pedidosya(att["raw_bytes"], att["filename"], att["msg_id"], etiqueta)
                        todos.extend(records)
                        stats["procesados"] += len(records)
                    except Exception as e:
                        stats["errores"].append(f"{att['filename']}: {e}")
                if atts and processed_label_id:
                    mark_as_processed(service, msg_id, processed_label_id)
            except Exception as e:
                stats["errores"].append(f"msg_id={msg_id}: {e}")
            time.sleep(PAUSE_S)
    try:
        stats["cargados"] = _load_to_bq(todos, bq_client)
    except Exception as e:
        stats["errores"].append(f"Load BQ: {e}")
    logger.info(f"[sync_pedidosya] procesados={stats['procesados']} cargados={stats['cargados']} errores={len(stats['errores'])}")
    return stats