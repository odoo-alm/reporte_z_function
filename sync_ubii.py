# =============================================================================
# sync_ubii.py  v4 — schema detalle por transacción (detalles_lote_U*.xlsx)
# Tabla BQ: mailer_raw.mailer_ubii_raw
# =============================================================================

import io
import json
import logging
import time
from datetime import datetime, timezone

import openpyxl
from google.cloud import bigquery
from utils import list_messages, get_attachments, mark_as_processed, circuit_is_open

logger = logging.getLogger(__name__)

BQ_PROJECT    = "gestion-365"
BQ_TABLE_UBII = "mailer_raw.mailer_ubii_raw"
PAUSE_S       = 5


def _parse_fecha(raw):
    """'01-04-2026' → '2026-04-01'"""
    if raw is None:
        return None
    if hasattr(raw, "strftime"):
        return raw.strftime("%Y-%m-%d")
    try:
        return datetime.strptime(str(raw).strip(), "%d-%m-%Y").strftime("%Y-%m-%d")
    except ValueError:
        return str(raw).strip()


def parse_ubii(raw_bytes, filename, msg_id, etiqueta):
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True)
    ws = wb[wb.sheetnames[0]]
    records = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or row[0] is None:
            continue
        try:
            records.append({
                "nro_lote":          int(row[0]) if row[0] is not None else None,
                "fecha":             _parse_fecha(row[1]),
                "hora":              str(row[2]).strip() if row[2] is not None else None,
                "tipo_transaccion":  str(row[3]).strip() if row[3] is not None else None,
                "referencia":        str(row[4]).strip() if row[4] is not None else None,
                "tarjeta":           str(row[5]).strip() if row[5] is not None else None,
                "tipo_tarjeta":      str(row[6]).strip() if row[6] is not None else None,
                "codigo_aprobacion": str(row[7]).strip() if row[7] is not None else None,
                "monto":             float(row[8]) if row[8] is not None else None,
                "total":             float(row[9]) if row[9] is not None else None,
                "monto_tasa":        str(row[10]).strip() if row[10] is not None else None,
                "monto_comision":    float(row[11]) if row[11] is not None else None,
                "tasa_islr":         str(row[12]).strip() if row[12] is not None else None,
                "monto_islr":        float(row[13]) if row[13] is not None else None,
                "reversada":         bool(row[14]) if row[14] is not None else None,
                "etiqueta":          etiqueta,
                "filename":          filename,
                "gmail_msg_id":      msg_id,
                "processed_at":      datetime.now(tz=timezone.utc).isoformat(),
            })
        except Exception as e:
            logger.warning(f"[sync_ubii] {filename}: fila {i} omitida — {e}")
    wb.close()
    logger.info(f"[sync_ubii] {filename}: {len(records)} filas.")
    return records

def _load_to_bq(records, bq_client):
    if not records:
        return 0
    buffer = io.BytesIO()
    for r in records:
        buffer.write((json.dumps(r, default=str, ensure_ascii=False) + "\n").encode())
    buffer.seek(0)
    job = bq_client.load_table_from_file(
        buffer, f"{BQ_PROJECT}.{BQ_TABLE_UBII}",
        job_config=bigquery.LoadJobConfig(
            source_format=bigquery.SourceFormat.NEWLINE_DELIMITED_JSON,
            write_disposition=bigquery.WriteDisposition.WRITE_APPEND,
            ignore_unknown_values=True,
        )
    )
    job.result()
    logger.info(f"[sync_ubii] {job.output_rows} filas cargadas.")
    return job.output_rows

def sync_ubii(service, bq_client, etiquetas, label_map=None, processed_label_id=None):
    stats = {"procesados": 0, "cargados": 0, "errores": []}
    todos = []
    for etiqueta in etiquetas:
        if circuit_is_open():
            stats["errores"].append("Circuit breaker abierto — abortado.")
            break
        label_id = (label_map or {}).get(etiqueta.upper())
        if not label_id:
            continue
        messages = list_messages(service, label_id)
        logger.info(f"[sync_ubii][{etiqueta}] {len(messages)} correos nuevos.")
        for msg_meta in messages:
            if circuit_is_open():
                stats["errores"].append("Circuit breaker abierto — abortado.")
                break
            msg_id = msg_meta["id"]
            try:
                atts = get_attachments(service, msg_id, keyword="detalles_lote", extensions=(".xlsx", ".xls"))
                for att in atts:
                    try:
                        records = parse_ubii(att["raw_bytes"], att["filename"], att["msg_id"], etiqueta)
                        todos.extend(records)
                        stats["procesados"] += len(records)
                    except Exception as e:
                        stats["errores"].append(f"{att['filename']}: {e}")
                if atts and processed_label_id:
                    mark_as_processed(service, msg_id, processed_label_id)
            except Exception as e:
                stats["errores"].append(f"msg_id={msg_id}: {e}")
            time.sleep(PAUSE_S)
    try:
        stats["cargados"] = _load_to_bq(todos, bq_client)
    except Exception as e:
        stats["errores"].append(f"Load BQ: {e}")
    logger.info(f"[sync_ubii] procesados={stats['procesados']} cargados={stats['cargados']} errores={len(stats['errores'])}")
    return stats