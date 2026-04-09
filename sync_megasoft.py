# =============================================================================
# sync_megasoft.py  v3 — usa utils.py
# =============================================================================

import io
import json
import logging
import time
from datetime import datetime, timezone

import openpyxl
from google.cloud import bigquery
from utils import list_messages, get_attachments, mark_as_processed, circuit_is_open

logger = logging.getLogger(__name__)

BQ_PROJECT    = "gestion-365"
BQ_TABLE_MEGA = "mailer_raw.mailer_megasoft_raw"
PAUSE_S       = 5

def parse_megasoft(raw_bytes, filename, msg_id, etiqueta):
    si = lambda v: int(v)   if v is not None else None
    sf = lambda v: float(v) if v is not None else None
    ss = lambda v: str(v).strip() if v is not None else None
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), read_only=True)
    ws = wb[wb.sheetnames[0]]
    records = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not header_found:
            if row[0] and str(row[0]).strip() == "Fecha Transacción":
                header_found = True
            continue
        if row[0] is None:
            continue
        try:
            records.append({
                "fecha_transaccion": ss(row[0]), "merchant_name": ss(row[1]),
                "afiliado": ss(row[2]), "caja": ss(row[3]), "terminal": ss(row[4]),
                "seq_num": si(row[5]), "pan_pci": ss(row[6]), "rif": ss(row[7]),
                "transaction_name": ss(row[8]), "acquirer_name": ss(row[9]),
                "card_name": ss(row[10]), "auth_id": ss(row[11]),
                "ref": ss(row[12]), "rrn": ss(row[13]),
                "fecha_cierre": ss(row[14]), "lote": ss(row[15]),
                "reversar": si(row[16]), "anulada": si(row[17]),
                "codigo_respuesta": ss(row[18]), "telefono": ss(row[19]),
                "monto": sf(row[20]), "tasa_cambio": sf(row[21]),
                "monto_cambio": sf(row[22]), "moneda_cambio": ss(row[23]),
                "etiqueta": etiqueta, "filename": filename,
                "gmail_msg_id": msg_id, "processed_at": datetime.now(tz=timezone.utc).isoformat(),
            })
        except Exception as e:
            logger.warning(f"[sync_megasoft] {filename}: fila omitida — {e}")
    wb.close()
    logger.info(f"[sync_megasoft] {filename}: {len(records)} filas.")
    return records

def _load_to_bq(records, bq_client):
    if not records:
        return 0
    buffer = io.BytesIO()
    for r in records:
        buffer.write((json.dumps(r, default=str, ensure_ascii=False) + "\n").encode())
    buffer.seek(0)
    job = bq_client.load_table_from_file(
        buffer, f"{BQ_PROJECT}.{BQ_TABLE_MEGA}",
        job_config=bigquery.LoadJobConfig(
            source_format=bigquery.SourceFormat.NEWLINE_DELIMITED_JSON,
            write_disposition=bigquery.WriteDisposition.WRITE_APPEND,
            ignore_unknown_values=True,
        )
    )
    job.result()
    logger.info(f"[sync_megasoft] {job.output_rows} filas cargadas.")
    return job.output_rows

def sync_megasoft(service, bq_client, etiquetas, label_map=None, processed_label_id=None):
    stats = {"procesados": 0, "cargados": 0, "errores": []}
    todos = []
    for etiqueta in etiquetas:
        if circuit_is_open():
            stats["errores"].append("Circuit breaker abierto — abortado.")
            break
        label_id = (label_map or {}).get(etiqueta.upper())
        if not label_id:
            continue
        messages = list_messages(service, label_id)
        logger.info(f"[sync_megasoft][{etiqueta}] {len(messages)} correos nuevos.")
        for msg_meta in messages:
            if circuit_is_open():
                stats["errores"].append("Circuit breaker abierto — abortado.")
                break
            msg_id = msg_meta["id"]
            try:
                atts = get_attachments(service, msg_id, keyword="TRANSACCIONES", extensions=(".xlsx", ".xls"))
                for att in atts:
                    try:
                        records = parse_megasoft(att["raw_bytes"], att["filename"], att["msg_id"], etiqueta)
                        todos.extend(records)
                        stats["procesados"] += len(records)
                    except Exception as e:
                        stats["errores"].append(f"{att['filename']}: {e}")
                if atts and processed_label_id:
                    mark_as_processed(service, msg_id, processed_label_id)
            except Exception as e:
                stats["errores"].append(f"msg_id={msg_id}: {e}")
            time.sleep(PAUSE_S)
    try:
        stats["cargados"] = _load_to_bq(todos, bq_client)
    except Exception as e:
        stats["errores"].append(f"Load BQ: {e}")
    logger.info(f"[sync_megasoft] procesados={stats['procesados']} cargados={stats['cargados']} errores={len(stats['errores'])}")
    return stats